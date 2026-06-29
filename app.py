from __future__ import annotations

import copy
import hmac
import json
import os
import threading
import uuid
from datetime import datetime
from functools import wraps
from pathlib import Path

from dotenv import load_dotenv
from flask import (
    Flask,
    abort,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

from aysa_bridge import claim_next_task, connector_status, heartbeat, submit_result
from query_engine import generate_pdf, run_properties

ROOT = Path(__file__).resolve().parent
load_dotenv(ROOT / ".env")
DATA_DIR = Path(os.getenv("DATA_DIR", ROOT / "data"))
REPORTS_DIR = Path(os.getenv("REPORTS_DIR", ROOT / "Reportes"))
EXCEL_PATH = DATA_DIR / "propiedades_rizzo.xlsx"
DATA_DIR.mkdir(parents=True, exist_ok=True)
REPORTS_DIR.mkdir(parents=True, exist_ok=True)
JOBS_DIR = DATA_DIR / "jobs"
JOBS_DIR.mkdir(parents=True, exist_ok=True)

app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "development-change-this")
app.config.update(
    MAX_CONTENT_LENGTH=12 * 1024 * 1024,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=os.getenv("COOKIE_SECURE", "false").lower() == "true",
)

JOBS: dict[str, dict] = {}
JOBS_LOCK = threading.Lock()
QUERY_LOCK = threading.Lock()

SERVICE_LABELS = {
    "tsg": "TSG Lanús",
    "aysa": "AySA",
    "edesur": "Edesur",
    "metrogas": "Metrogas",
    "arba": "ARBA",
    "expensas": "Expensas",
}


def _job_path(job_id: str) -> Path:
    return JOBS_DIR / f"{job_id}.json"


def _write_job_snapshot(job_id: str, payload: dict) -> None:
    try:
        path = _job_path(job_id)
        temporary = path.with_suffix(".tmp")
        temporary.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        temporary.replace(path)
    except Exception:
        # El guardado parcial nunca debe interrumpir una consulta real.
        pass


def _save_job_snapshot(job_id: str) -> None:
    with JOBS_LOCK:
        job = JOBS.get(job_id)
        snapshot = copy.deepcopy(job) if job else None
    if snapshot is not None:
        _write_job_snapshot(job_id, snapshot)


def _load_job_snapshot(job_id: str) -> dict | None:
    path = _job_path(job_id)
    if not path.exists():
        return None
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
        return payload if isinstance(payload, dict) else None
    except Exception:
        return None


def _restore_saved_jobs() -> None:
    for path in JOBS_DIR.glob("*.json"):
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
            if not isinstance(payload, dict) or not payload.get("id"):
                continue
            if payload.get("status") in {"queued", "running"}:
                payload["status"] = "interrupted"
                payload["error"] = (
                    "La consulta se interrumpió por un reinicio del servidor. "
                    "Los resultados terminados quedaron guardados en esta pantalla."
                )
            JOBS[str(payload["id"])] = payload
        except Exception:
            continue


def _connector_authorized() -> bool:
    expected = os.getenv("CONNECTOR_TOKEN", "").strip()
    if not expected:
        return False
    authorization = request.headers.get("Authorization", "")
    supplied = authorization[7:].strip() if authorization.startswith("Bearer ") else ""
    return _safe_equal(supplied, expected)




def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login", next=request.path))
        return view(*args, **kwargs)
    return wrapped


def _safe_equal(left: str, right: str) -> bool:
    return hmac.compare_digest((left or "").encode(), (right or "").encode())


_restore_saved_jobs()


def load_properties() -> list[dict]:
    if not EXCEL_PATH.exists():
        return []
    workbook = load_workbook(EXCEL_PATH, data_only=True, read_only=True)
    sheet = workbook["Propiedades"]
    properties: list[dict] = []
    for excel_row, row in enumerate(sheet.iter_rows(min_row=4, values_only=True), start=4):
        values = list(row) + [None] * 25
        if not values[0]:
            continue
        properties.append({
            "id": str(excel_row),
            "propietario": str(values[0] or "").strip(),
            "inquilino": str(values[1] or "").strip(),
            "direccion": str(values[2] or "").strip(),
            "inicio": values[3],
            "padron": str(values[4] or "").strip(),
            "aysa": str(values[5] or "").strip(),
            "edesur": str(values[6] or "").strip(),
            "medidor": str(values[7] or "").strip(),
            "metrogas": str(values[8] or "").strip(),
            "link_expensas": str(values[9] or "").strip(),
            "usuario_expensas": str(values[10] or "").strip(),
            "pass_expensas": str(values[11] or "").strip(),
            "partido": str(values[12] or "").strip(),
            "partida": str(values[13] or "").strip(),
            "observaciones": str(values[14] or "").strip(),
        })
    workbook.close()
    return properties


def _account_reference(prop: dict, service: str) -> str:
    if service == "tsg":
        return prop.get("padron") or "—"
    if service == "aysa":
        return prop.get("aysa") or "—"
    if service == "edesur":
        client = prop.get("edesur") or "—"
        meter = prop.get("medidor") or "—"
        return f"{client} / {meter}"
    if service == "metrogas":
        return prop.get("metrogas") or "—"
    if service == "arba":
        return f"{prop.get('partido') or '—'}-{prop.get('partida') or '—'}"
    if service == "expensas":
        return "Link cargado" if prop.get("link_expensas") else "—"
    return "—"


def _money_display(value: object) -> str:
    if value in (None, "", "—", "-"):
        return "—"
    try:
        raw = str(value).strip().replace("$", "").replace(" ", "")
        if "," in raw:
            raw = raw.replace(".", "").replace(",", ".")
        number = float(raw)
        return f"$ {number:,.0f}".replace(",", ".")
    except Exception:
        return str(value)


def _initial_live_results(properties: list[dict], services: list[str]) -> list[dict]:
    items: list[dict] = []
    for prop in properties:
        items.append({
            "id": str(prop.get("id") or ""),
            "propietario": prop.get("propietario") or "Sin propietario",
            "inquilino": prop.get("inquilino") or "",
            "direccion": prop.get("direccion") or "",
            "status": "pending",
            "services": [
                {
                    "key": key,
                    "label": SERVICE_LABELS[key],
                    "account": _account_reference(prop, key),
                    "status": "pending",
                    "state": "Pendiente",
                    "debt": "—",
                    "observations": "",
                }
                for key in services
            ],
        })
    return items


def report_inventory() -> list[dict]:
    reports: list[dict] = []
    for path in REPORTS_DIR.rglob("*.pdf"):
        if not path.is_file():
            continue
        relative = path.relative_to(REPORTS_DIR)
        stat = path.stat()
        reports.append({
            "id": relative.as_posix(),
            "name": path.stem.replace("_", " "),
            "filename": path.name,
            "service": relative.parts[0] if len(relative.parts) > 1 else "General",
            "modified": datetime.fromtimestamp(stat.st_mtime),
            "size_kb": max(1, round(stat.st_size / 1024)),
        })
    reports.sort(key=lambda item: item["modified"], reverse=True)
    return reports


def safe_report_path(report_id: str) -> Path:
    candidate = (REPORTS_DIR / report_id).resolve()
    root = REPORTS_DIR.resolve()
    if root != candidate and root not in candidate.parents:
        abort(404)
    if candidate.suffix.lower() != ".pdf" or not candidate.is_file():
        abort(404)
    return candidate


@app.context_processor
def inject_globals():
    return {"current_year": datetime.now().year}


@app.get("/health")
def health():
    return {"status": "ok"}


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        expected_user = os.getenv("APP_USERNAME", "")
        expected_pass = os.getenv("APP_PASSWORD", "")
        if not expected_user or not expected_pass:
            flash("El administrador todavía no configuró el usuario y la contraseña.", "danger")
        elif _safe_equal(request.form.get("username", ""), expected_user) and _safe_equal(
            request.form.get("password", ""), expected_pass
        ):
            session.clear()
            session["logged_in"] = True
            session.permanent = True
            return redirect(request.args.get("next") or url_for("query_page"))
        else:
            flash("Usuario o contraseña incorrectos.", "danger")
    return render_template("login.html")


@app.post("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.get("/")
@login_required
def dashboard():
    # Al abrir el sistema, ir directamente a la pantalla de consultas.
    return redirect(url_for("query_page"))


@app.get("/propiedades")
@login_required
def properties_page():
    query = request.args.get("q", "").strip().lower()
    properties = load_properties()
    if query:
        properties = [
            prop for prop in properties
            if query in " ".join((prop["propietario"], prop["inquilino"], prop["direccion"])).lower()
        ]
    return render_template("properties.html", properties=properties, query=request.args.get("q", ""))


@app.get("/reportes")
@login_required
def reports_page():
    service = request.args.get("servicio", "").strip()
    reports = report_inventory()
    services = sorted({report["service"] for report in reports})
    if service:
        reports = [report for report in reports if report["service"] == service]
    return render_template("reports.html", reports=reports, services=services, selected_service=service)


@app.get("/reportes/ver")
@login_required
def view_report():
    report_id = request.args.get("id", "")
    path = safe_report_path(report_id)
    return render_template("report_view.html", report_id=report_id, report_name=path.stem.replace("_", " "))


@app.get("/reportes/archivo")
@login_required
def report_file():
    path = safe_report_path(request.args.get("id", ""))
    download = request.args.get("download") == "1"
    return send_file(path, mimetype="application/pdf", as_attachment=download, download_name=path.name)


@app.route("/excel", methods=["GET", "POST"])
@login_required
def excel_page():
    if request.method == "POST":
        upload = request.files.get("excel")
        if not upload or not upload.filename:
            flash("Seleccioná un archivo Excel.", "warning")
            return redirect(url_for("excel_page"))
        filename = secure_filename(upload.filename)
        if not filename.lower().endswith(".xlsx"):
            flash("El archivo debe ser .xlsx", "danger")
            return redirect(url_for("excel_page"))
        temporary = DATA_DIR / "propiedades_rizzo_nuevo.xlsx"
        upload.save(temporary)
        try:
            workbook = load_workbook(temporary, read_only=True)
            if "Propiedades" not in workbook.sheetnames:
                raise ValueError("No existe la hoja Propiedades")
            workbook.close()
            if EXCEL_PATH.exists():
                backup = DATA_DIR / f"propiedades_rizzo_backup_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
                EXCEL_PATH.replace(backup)
            temporary.replace(EXCEL_PATH)
            flash("El Excel fue actualizado correctamente.", "success")
        except Exception as exc:
            temporary.unlink(missing_ok=True)
            flash(f"No se pudo usar ese Excel: {exc}", "danger")
        return redirect(url_for("excel_page"))
    return render_template(
        "excel.html",
        excel_exists=EXCEL_PATH.exists(),
        excel_modified=(datetime.fromtimestamp(EXCEL_PATH.stat().st_mtime) if EXCEL_PATH.exists() else None),
        property_count=len(load_properties()),
    )


@app.get("/excel/descargar")
@login_required
def download_excel():
    if not EXCEL_PATH.exists():
        abort(404)
    return send_file(EXCEL_PATH, as_attachment=True, download_name="propiedades_rizzo.xlsx")


@app.get("/consultar")
@login_required
def query_page():
    return render_template("query.html", properties=load_properties())


def _update_job(job_id: str, **changes):
    with JOBS_LOCK:
        if job_id in JOBS:
            JOBS[job_id].update(changes)
    _save_job_snapshot(job_id)


def _append_log(job_id: str, line: str):
    with JOBS_LOCK:
        if job_id in JOBS:
            JOBS[job_id]["logs"].append(line)
            JOBS[job_id]["logs"] = JOBS[job_id]["logs"][-500:]
    _save_job_snapshot(job_id)


def _process_live_event(job_id: str, event: dict) -> None:
    changed = False
    with JOBS_LOCK:
        job = JOBS.get(job_id)
        if not job:
            return
        property_id = str(event.get("property_id") or "")
        item = next((row for row in job.get("live_results", []) if row.get("id") == property_id), None)
        if not item:
            return

        event_type = event.get("type")
        if event_type == "property_started":
            item["status"] = "running"
            job["current_property"] = item.get("propietario") or "Propiedad"
            changed = True
        elif event_type == "property_completed":
            item["status"] = "done"
            job["properties_done"] = int(event.get("position") or 0)
            changed = True
        else:
            service_key = str(event.get("service_key") or "")
            service = next((row for row in item.get("services", []) if row.get("key") == service_key), None)
            if not service:
                return
            if event_type == "service_started":
                service.update(status="running", state="Consultando…", debt="—", observations="")
                job["current_service"] = service.get("label") or "Servicio"
                changed = True
            elif event_type == "service_waiting":
                service.update(
                    status="running",
                    state="Pendiente de conector local",
                    debt="—",
                    observations="Abrí el Conector AySA en la computadora de la inmobiliaria",
                )
                job["current_service"] = "AySA · esperando conector local"
                changed = True
            elif event_type == "service_completed":
                result = event.get("result") if isinstance(event.get("result"), dict) else {}
                state = str(result.get("estado") or "Sin estado")
                debt = result.get("deuda")
                if service_key == "tsg":
                    debt = result.get("vencida") if result.get("vencida") is not None else result.get("total")
                observations = result.get("observaciones") or result.get("obs") or ""
                service.update(
                    status="error" if "error" in state.lower() else "done",
                    state=state,
                    debt=_money_display(debt),
                    observations=str(observations)[:300],
                )
                completed = sum(
                    1
                    for prop in job.get("live_results", [])
                    for row in prop.get("services", [])
                    if row.get("status") in {"done", "error"}
                )
                job["done"] = completed
                job["percent"] = round(completed * 100 / max(int(job.get("total") or 1), 1))
                changed = True
    if changed:
        _save_job_snapshot(job_id)


def _run_job(job_id: str, selected: list[dict], services: list[str]):
    if not QUERY_LOCK.acquire(blocking=False):
        _update_job(job_id, status="error", error="Ya hay otra consulta ejecutándose.")
        return
    try:
        _update_job(job_id, status="running", started_at=datetime.now().isoformat())

        def progress(done: int, total: int):
            _update_job(job_id, properties_done=done, properties_total=total)

        results = run_properties(
            selected,
            services,
            lambda line: _append_log(job_id, line),
            progress,
            lambda event: _process_live_event(job_id, event),
        )
        pdf = generate_pdf(results)
        report_id = pdf.relative_to(REPORTS_DIR).as_posix()
        _append_log(job_id, f"\nReporte generado: {pdf.name}")
        _update_job(
            job_id,
            status="done",
            finished_at=datetime.now().isoformat(),
            percent=100,
            report_id=report_id,
        )
    except Exception as exc:
        _append_log(job_id, f"Error general: {exc}")
        _update_job(job_id, status="error", error=str(exc))
    finally:
        QUERY_LOCK.release()


@app.post("/api/jobs")
@login_required
def create_job():
    payload = request.get_json(silent=True) or {}
    property_ids = {str(item) for item in payload.get("property_ids", [])}
    services = [str(item) for item in payload.get("services", []) if str(item) in {
        "tsg", "aysa", "edesur", "metrogas", "arba", "expensas"
    }]
    # AySA se procesa al final porque depende del conector local de Windows.
    if "aysa" in services:
        services = [service for service in services if service != "aysa"] + ["aysa"]
    if not property_ids:
        return jsonify({"error": "Seleccioná al menos una propiedad."}), 400
    if not services:
        return jsonify({"error": "Seleccioná al menos un servicio."}), 400
    properties = [prop for prop in load_properties() if prop["id"] in property_ids]
    if not properties:
        return jsonify({"error": "No se encontraron las propiedades seleccionadas."}), 400
    with JOBS_LOCK:
        if any(job.get("status") in {"queued", "running"} for job in JOBS.values()):
            return jsonify({"error": "Ya hay una consulta en curso. Esperá a que finalice."}), 409
        job_id = uuid.uuid4().hex[:12]
        JOBS[job_id] = {
            "id": job_id,
            "status": "queued",
            "created_at": datetime.now().isoformat(),
            "done": 0,
            "total": len(properties) * len(services),
            "properties_done": 0,
            "properties_total": len(properties),
            "percent": 0,
            "current_property": "",
            "current_service": "",
            "live_results": _initial_live_results(properties, services),
            "logs": [],
            "report_id": None,
            "error": None,
        }
    _save_job_snapshot(job_id)
    thread = threading.Thread(target=_run_job, args=(job_id, properties, services), daemon=True)
    thread.start()
    return jsonify({"job_id": job_id}), 202


@app.get("/api/jobs/<job_id>")
@login_required
def get_job(job_id: str):
    with JOBS_LOCK:
        job = JOBS.get(job_id)
    if not job:
        job = _load_job_snapshot(job_id)
        if job:
            with JOBS_LOCK:
                JOBS[job_id] = job
    if not job:
        abort(404)
    return jsonify(job)


@app.get("/api/connector/aysa/status")
@login_required
def aysa_connector_status():
    return jsonify(connector_status())


@app.post("/api/connector/aysa/heartbeat")
def aysa_connector_heartbeat():
    if not _connector_authorized():
        return jsonify({"error": "Token de conector inválido o no configurado."}), 401
    heartbeat()
    return jsonify({"ok": True})


@app.get("/api/connector/aysa/next")
def aysa_connector_next():
    if not _connector_authorized():
        return jsonify({"error": "Token de conector inválido o no configurado."}), 401
    task = claim_next_task()
    if not task:
        return ("", 204)
    return jsonify(task)


@app.post("/api/connector/aysa/result/<task_id>")
def aysa_connector_result(task_id: str):
    if not _connector_authorized():
        return jsonify({"error": "Token de conector inválido o no configurado."}), 401
    payload = request.get_json(silent=True) or {}
    result = payload.get("result")
    if not isinstance(result, dict):
        return jsonify({"error": "Resultado inválido."}), 400
    if not submit_result(task_id, result):
        return jsonify({"error": "La tarea no existe o ya expiró."}), 404
    return jsonify({"ok": True})


@app.errorhandler(413)
def too_large(_error):
    flash("El archivo supera el límite permitido de 12 MB.", "danger")
    return redirect(url_for("excel_page"))


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "5000")), debug=False, threaded=True)
